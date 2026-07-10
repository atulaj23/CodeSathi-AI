from pypdf import PdfReader
from docx import Document
from openpyxl import load_workbook
import os



def read_pdf(file_path):

    try:
        reader = PdfReader(file_path)

        text = ""

        for page in reader.pages:
            text += page.extract_text() or ""

        return text

    except Exception as e:
        print("PDF ERROR:", e)
        return ""




def read_docx(file_path):

    try:

        doc = Document(file_path)

        text = ""

        for paragraph in doc.paragraphs:
            text += paragraph.text + "\n"

        return text

    except Exception as e:
        print("DOCX ERROR:", e)
        return ""




def read_txt(file_path):

    try:

        with open(
            file_path,
            "r",
            encoding="utf-8"
        ) as file:

            return file.read()

    except Exception as e:
        print("TXT ERROR:", e)
        return ""




def read_excel(file_path):

    try:

        workbook = load_workbook(
            file_path,
            data_only=True
        )

        text = ""

        for sheet in workbook.sheetnames:

            ws = workbook[sheet]

            text += f"\nSheet: {sheet}\n"

            for row in ws.iter_rows(values_only=True):

                text += " ".join(
                    [
                        str(cell)
                        for cell in row
                        if cell is not None
                    ]
                )

                text += "\n"

        return text

    except Exception as e:
        print("EXCEL ERROR:", e)
        return ""




def read_file(file_path):

    if not file_path:
        return ""

    extension = os.path.splitext(
        file_path
    )[1].lower()



    if extension == ".pdf":

        return read_pdf(file_path)



    elif extension == ".docx":

        return read_docx(file_path)



    elif extension == ".txt":

        return read_txt(file_path)



    elif extension == ".xlsx":

        return read_excel(file_path)



    else:

        return "Unsupported file type"